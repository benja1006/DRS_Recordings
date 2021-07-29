"""The main file."""

import os
import openpyxl
from shutil import copyfile
from openpyxl.styles import PatternFill
import filecmp
fill = PatternFill("solid", fgColor='92d051')


def findFiles(date, env):
    """Return desired files.

    Check file names against list of phone numbers from excel sheet. Return the
    overlap.
    """
    source = env['SOURCE']
    phoneNums = findPhoneNums(env)
    # print(phoneNums)
    returnFiles = []
    if(not os.path.isdir(os.path.join(source, date))):
        return False

    files = os.listdir(os.path.join(source, date))
    # print(files)
    # print(phoneNums)
    doneFiles = []
    for filename in files:
        for phoneNum in phoneNums:
            # if phoneNum in filename and 'Completed Enrollment' in filename:
            if phoneNum in filename:
                returnFiles.append((filename, phoneNum))
                # files.remove(filename)
                doneFiles.append(filename)
                break
    files = set(files) - set(doneFiles)
    print('outputting ' + str(len(returnFiles)) + ' files')
    """ returnFiles should now contain all files where the phone number exists
    in the excel file
    now find files that are labeled as completed enrollment that aren't from a
    known phone number
    """
    manualFiles = []
    for file in files:
        if 'Completed Enrollment' in file:
            substr = file.split('_')[3]
            num = substr[0:10]
            manualFiles.append((file, num))
    # returnFiles consists of all files with phone numbers from the spreadsheet
    # manualFiles consists of files with phone numbers not in the spreadsheet
    # print('Check on the following files: ')
    # for file in manualFiles:
    #    print(str(file[0]), ' from ', str(file[1]))
    returnArr = (returnFiles, manualFiles)
    return returnArr


def findPhoneNums(env):
    """Return list of enrolled phone numbers that haven't been checked off."""
    wbPath = os.path.abspath(env['EXCEL'])
    # print('wbPath=', wbPath)
    wb = openpyxl.load_workbook(wbPath)
    ws = wb.active
    i = 2
    phoneNums = []
    while ws['A'+str(i)].value is not None:
        if ws['A'+str(i)].fill == fill:
            i += 1
            continue
        num = ws['S' + str(i)].value
        if num is not None:
            phoneNums.append(num)
        i += 1
    return phoneNums


def ssLookup(phoneNum, env):
    """Look up phone number in spreadsheet and returns user."""
    wbPath = os.path.abspath(env['EXCEL'])
    wb = openpyxl.load_workbook(wbPath)
    ws = wb.active
    i = 1
    while ws['A'+str(i)].value is not None:
        if ws['S'+str(i)].value == phoneNum:
            date = str(ws['L' + str(i)].value)  # date in YYYY-MM-DD HH:MM:SS
            print(date)
            date = date[0:10]  # date in YYYY-MM-DD need MM-DD-M_D_YYYY
            dYear = date[0:4]
            dMonth = date[5:7]
            dDate = date[8:10]
            newDate = dMonth + '-' + dDate + '-' + dYear
            user = {
                "FNAME": str(ws['G'+str(i)].value),
                "LNAME": str(ws['F'+str(i)].value),
                "Enrollment Date": newDate
            }
            break
        i += 1
    return user


def markExcelRow(phoneNum, env):
    """Highlight the row and resave the xlsx file."""
    wbPath = os.path.abspath(env['EXCEL'])
    wb = openpyxl.load_workbook(wbPath)
    ws = wb.active
    i = 1
    while ws['A'+str(i)].value is not None:
        if ws['S'+str(i)].value == phoneNum:
            for cell in ws[str(i)]:
                cell.fill = fill
            ws['AH'+str(i)].value = 'Yes'
            wb.save(wbPath)
            return
        else:
            i += 1


def renameFile(fileArr, date, env, repeat):
    """Copy .wav to new location and changes name."""
    file = os.path.abspath(os.path.join(env['SOURCE'], date, fileArr[0]))
    user = ssLookup(fileArr[1], env)
    if(not os.path.isdir(os.path.join(env['DEST'], date))):
        os.mkdir(os.path.join(env['DEST'], date))
    # print(user['Enrollment Date'])
    if(repeat == 0):
        repeatStr = ''
    else:
        repeatStr = str(repeat)
    destination = os.path.abspath(os.path.join(env['DEST'], date,
                                               user['FNAME'] + ' ' +
                                               user['LNAME'] + ' ' +
                                               user['Enrollment Date'] +
                                               repeatStr + '.wav'))
    # check to make sure file with same name doesn't already exist
    if(os.path.isfile(destination)):
        # if it does see if it is the same file
        if(filecmp.cmp(file, destination)):
            # return True meaning day has been run
            return True
        else:
            # rerun function with repeat incremented by 1
            return renameFile(fileArr, date, env, repeat + 1)
    copyfile(file, destination)
    return False


def main():
    """Perform the main function."""
    print("Please enter the date that you wish to lookup in M_D_YYYY, using"
          "two digits for month and date when the first digit isn't 0")
    date = input()
    filesArr = findFiles(date)
    # fileArr is a tuple (fileName, phoneNum)
    for fileArr in filesArr:
        renameFile(fileArr, date)
        print(fileArr[1])
        markExcelRow(fileArr[1])


if __name__ == '__main__':
    main()
